import requests
import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from termcolor import colored

# Define the base url of the api
base_url = "https://api.kampusmerdeka.kemdikbud.go.id/magang/browse/position"

# Define the parameters for the api
params = {
    "offset": 0,
    "limit": 20,
    "sort_by": "published_time",
    "order": "desc"
}

# Define the maximum offset value
max_offset = 5513

# Create an empty list to store all data
all_data = []

# Loop through the offset values with a step of 20
for offset in range(0, max_offset + 1, 20):
    # Update the offset parameter
    params["offset"] = offset

    # Make a get request to the api
    response = requests.get(base_url, params=params)

    # Check if the response is successful
    if response.status_code == 200:
        # Get the json data from the response
        data = response.json()

        # Add the data to the all_data list
        all_data.extend(data['data'])

        # Print a message to indicate the progress
        print(colored(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] -", "green"), f"Data fetched for offset {offset}")
    else:
        # Print an error message if the response is not successful
        print(colored(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] -", "red"), f"Error: {response.status_code}")

# Convert the list of data to a DataFrame
df = pd.DataFrame(all_data)

# Add a "No" column at the beginning of the DataFrame
df.insert(0, 'no', range(1, 1 + len(df)))

# Save the DataFrame to an Excel file
df.to_excel('scrape-magang.xlsx', index=False)

# Print a message to indicate the progress
print(colored(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] -", "green"), "Data saved to Excel file")

# Load the workbook
book = load_workbook('scrape-magang.xlsx')

# Select the default sheet
sheet = book.active

# Define a fill color
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Apply the fill color to the header row
for cell in sheet[1]:
    cell.fill = fill

# Save the workbook
book.save('scrape-magang.xlsx')

# Print a message to indicate the progress
print(colored(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] -", "green"), "Workbook saved with colored header")